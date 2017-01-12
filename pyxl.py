# # To read an xlsx file
# from openpyxl import load_workbook
# wb = load_workbook(filename = 'loc not found.xlsx')
# sheet_ranges = wb['Sheet1']
# for row in sheet_ranges:
# 	for cell in row:
# 		obj =cell.value
# 		print obj
# import csv
# import re
#
# file_name = "all_delivered.csv"
#
#
# def get_csv_records(file_name):
# 	data = open(file_name)
# 	reader = csv.reader(data)
# 	records_list = []
# 	keys = reader.next()
#
# 	for row in reader:
# 		record_dict = dict(zip(keys, row))
# 		records_list.append(record_dict)
# 	data.close()
#
# 	return records_list
#
#
# def get_pattern(file_data):
# 	format_freq_dist = {}
#
# 	for x in file_data:
# 		new_id = re.sub(r'[a-zA-Z]', "A", x['address'])
# 		new_id1 = re.sub(r'[0-9]', "9", new_id)
#
# 		format_freq_dist.setdefault(new_id1, 0)
# 		format_freq_dist[new_id1] += 1
# 	return format_freq_dist
#
#
# def get_format(file_data):
# 	value_freq_dist = {}
#
# 	for i in file_data:
# 		for j in i:
# 			value_freq_dist.setdefault(i[j], 0)
# 			value_freq_dist[i[j]] += 1
# 	return value_freq_dist
#
#
# # def writeToCSV(filename, contents_to_write, keys):
# # 	# PURPOSE:
# # 	#       This function reads the contents of a List of Dictionaries,
# # 	#       and stores them in a csv file.
# # 	with open(filename, 'wb') as f:
# # 		w = csv.DictWriter(f, keys)
# # 		w.writeheader()
# # 		for row in contents_to_write:
# # 			w.writerow(row)
# # 	return True
#
#
# if __name__ == '__main__':
# 	file_data = get_csv_records(file_name)
# 	myresults = get_pattern(file_data)
# 	finalresult = get_format(file_data)
# 	final_list = []
# 	final_list1 = []
# 	for item in myresults:
# 		d1 = {
# 			'Format': item,
# 			'Occurances': myresults[item],
# 		}
# 		final_list.append(d1)
# 	for item in finalresult:
# 		d2 = {
# 			'Value': item,
# 			'Occurances': finalresult[item]
# 		}
# 		final_list1.append(d2)
# 	cardinality = (len(final_list) / float(len(file_data)) * 100)
# 	cardinality1 = (len(final_list1)) / float(len(file_data) * 100)
# 	print "cardinality of pattern is %s" % cardinality
# 	print "cardinality of value is %s" % cardinality1

# writeToCSV("all_format_add.csv",final_list, ['Format', 'Occurances'])
# writeToCSV("all_value_add.csv",final_list1, ['Value', 'Occurances'])
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
     for col in range(27, 54):
         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename = dest_filename)
